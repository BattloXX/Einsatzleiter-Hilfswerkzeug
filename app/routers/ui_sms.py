"""Admin-UI: SMS-Gruppen, Einsatzinfo-SMS-Konfiguration und manueller SMS-Versand."""
from __future__ import annotations

import logging
from datetime import UTC, datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.orm import Session

from app.core.audit import write_audit
from app.core.permissions import has_role, require_role
from app.core.templating import templates
from app.db import get_db
from app.models.master import AlarmType, Member, OrgSettings
from app.models.sms import SmsEinsatzinfoRecipient, SmsGroup, SmsGroupMember, SmsLog

router = APIRouter(prefix="/admin")
logger = logging.getLogger("einsatzleiter.ui_sms")


# ── Hilfsfunktionen ───────────────────────────────────────────────────────────

def _require_org(user) -> int:
    """Gibt org_id des Users zurueck. Wirft 403 wenn kein Org-Kontext vorhanden."""
    if not user.org_id:
        raise HTTPException(status_code=403, detail="Kein Org-Kontext")
    return user.org_id


def _sms_groups_for_org(db: Session, org_id: int) -> list[SmsGroup]:
    return (
        db.query(SmsGroup)
        .filter(SmsGroup.org_id == org_id)
        .order_by(SmsGroup.display_order, SmsGroup.name)
        .all()
    )


def _active_members(db: Session, org_id: int) -> list[Member]:
    return (
        db.query(Member)
        .filter(Member.org_id == org_id, Member.active.is_(True))
        .order_by(Member.lastname, Member.firstname)
        .all()
    )


# ── SMS-Gruppen ───────────────────────────────────────────────────────────────

@router.get("/gruppen", response_class=HTMLResponse)
async def sms_groups_page(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    user = request.state.user
    org_id = _require_org(user)
    groups = _sms_groups_for_org(db, org_id)
    members = _active_members(db, org_id)
    return templates.TemplateResponse(request, "admin/sms_groups.html", {
        "user": user,
        "groups": groups,
        "members": members,
        "saved": request.query_params.get("saved"),
        "error": request.query_params.get("error"),
    })


@router.post("/gruppen/neu")
async def sms_group_create(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    user = request.state.user
    org_id = _require_org(user)
    name = name.strip()
    if not name:
        return RedirectResponse("/admin/sms-gruppen?error=empty", status_code=303)
    grp = SmsGroup(
        org_id=org_id,
        name=name,
        description=description.strip() or None,
        display_order=0,
        created_at=datetime.now(UTC),
    )
    db.add(grp)
    write_audit(db, "admin.sms_group.created", org_id=org_id, user_id=user.id,
                entity_type="sms_group", payload={"name": name})
    db.commit()
    return RedirectResponse(f"/admin/gruppen?saved=1#gruppe-{grp.id}", status_code=303)


@router.post("/gruppen/{group_id}/edit")
async def sms_group_edit(
    group_id: int,
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    user = request.state.user
    org_id = _require_org(user)
    grp = db.get(SmsGroup, group_id)
    if not grp or grp.org_id != org_id:
        raise HTTPException(status_code=404)
    grp.name = name.strip() or grp.name
    grp.description = description.strip() or None
    write_audit(db, "admin.sms_group.edited", org_id=org_id, user_id=user.id,
                entity_type="sms_group", entity_id=group_id)
    db.commit()
    return RedirectResponse(f"/admin/gruppen?saved=1#gruppe-{group_id}", status_code=303)


@router.post("/gruppen/{group_id}/loeschen")
async def sms_group_delete(
    group_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    user = request.state.user
    org_id = _require_org(user)
    grp = db.get(SmsGroup, group_id)
    if not grp or grp.org_id != org_id:
        raise HTTPException(status_code=404)
    db.delete(grp)
    write_audit(db, "admin.sms_group.deleted", org_id=org_id, user_id=user.id,
                entity_type="sms_group", entity_id=group_id)
    db.commit()
    return RedirectResponse("/admin/gruppen?saved=1", status_code=303)


@router.post("/gruppen/{group_id}/mitglieder")
async def sms_group_set_members(
    group_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Setzt die Mitglieder einer Gruppe (vollstaendiger Ersatz via HTMX-Formular)."""
    user = request.state.user
    org_id = _require_org(user)
    grp = db.get(SmsGroup, group_id)
    if not grp or grp.org_id != org_id:
        raise HTTPException(status_code=404)

    form = await request.form()
    selected_ids = {int(v) for k, v in form.multi_items() if k == "member_id"}

    # Alle bestehenden Eintraege loeschen und neu anlegen
    db.query(SmsGroupMember).filter(SmsGroupMember.sms_group_id == group_id).delete()
    for mid in selected_ids:
        db.add(SmsGroupMember(sms_group_id=group_id, member_id=mid))

    write_audit(db, "admin.sms_group.members_updated", org_id=org_id, user_id=user.id,
                entity_type="sms_group", entity_id=group_id,
                payload={"count": len(selected_ids)})
    db.commit()

    # HTMX-Partial: Gruppen-Liste neu rendern
    groups = _sms_groups_for_org(db, org_id)
    members = _active_members(db, org_id)
    return templates.TemplateResponse(request, "admin/sms_groups.html", {
        "user": user,
        "groups": groups,
        "members": members,
        "saved": "1",
        "error": None,
    })


# ── Gruppen Excel-Import ──────────────────────────────────────────────────────

@router.get("/gruppen/excel-import")
def gruppen_excel_import_redirect():
    return RedirectResponse("/admin/gruppen", status_code=303)


@router.post("/gruppen/excel-import")
async def import_gruppen_excel(
    request: Request,
    file: UploadFile = File(...),
    target_group_id: int = Form(0),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Massenimport von Gruppen und Mitgliedschaften aus Excel.

    Erwartete Spalten:
      Gruppenname (required wenn keine Zielgruppe gewaehlt), Nachname (required),
      Vorname (required), Telefon (optional), E-Mail (optional).

    Personen werden aus den Mitgliedern gesucht; fehlende werden als Mitglieder angelegt.
    Wenn target_group_id gesetzt: alle Zeilen dieser Gruppe zuordnen (keine Gruppenname-Spalte noetig).
    """
    import io as _io
    import urllib.parse
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/admin/gruppen?error=openpyxl_missing", status_code=303)

    raw = await file.read()
    if not raw:
        return RedirectResponse("/admin/gruppen?error=empty_file", status_code=303)

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return RedirectResponse("/admin/gruppen?error=invalid_excel", status_code=303)

    _GROUP_ALIASES     = {"gruppenname", "gruppe", "group", "grp"}
    _LASTNAME_ALIASES  = {"nachname", "lastname", "name", "zuname", "familienname"}
    _FIRSTNAME_ALIASES = {"vorname", "firstname", "rufname"}
    _PHONE_ALIASES     = {"telefon", "phone", "tel", "mobil", "handy", "mobiltelefon", "telefonnummer"}
    _EMAIL_ALIASES     = {"e-mail", "email", "mail"}

    try:
        header_row = next(ws.iter_rows(max_row=1))
    except StopIteration:
        return RedirectResponse("/admin/gruppen?error=empty_sheet", status_code=303)
    headers = [str(c.value or "").strip().lower() for c in header_row]
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in _GROUP_ALIASES:
            col_map.setdefault("group", i)
        elif h in _LASTNAME_ALIASES:
            col_map.setdefault("lastname", i)
        elif h in _FIRSTNAME_ALIASES:
            col_map.setdefault("firstname", i)
        elif h in _PHONE_ALIASES:
            col_map.setdefault("phone", i)
        elif h in _EMAIL_ALIASES:
            col_map.setdefault("email", i)

    # Wenn Zielgruppe gewaehlt, ist Gruppenname-Spalte nicht Pflicht
    required_cols = ("lastname", "firstname") if target_group_id else ("group", "lastname", "firstname")
    missing = [k for k in required_cols if k not in col_map]
    if missing:
        found = ", ".join('"' + h + '"' for h in headers[:8] if h)
        detail = "Gefundene Spalten: " + found + ". Erwartet: Zuname/Nachname und Vorname" + (
            "" if target_group_id else " sowie Gruppenname (oder Zielgruppe im Formular waehlen)"
        ) + "."
        return RedirectResponse(
            f"/admin/gruppen?error=missing_columns&error_detail={urllib.parse.quote(detail)}",
            status_code=303,
        )

    user = request.state.user
    org_id = _require_org(user)

    # Zielgruppe vorab laden (wenn angegeben)
    target_group: SmsGroup | None = None
    if target_group_id:
        target_group = db.get(SmsGroup, target_group_id)
        if not target_group or target_group.org_id != org_id:
            return RedirectResponse("/admin/gruppen?error=invalid_group", status_code=303)

    group_cache: dict[str, SmsGroup] = {}
    if target_group:
        group_cache[target_group.name] = target_group
    members_created = 0
    members_updated = 0
    groups_created = 0
    memberships_added = 0
    skipped = 0
    row_errors = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        sp = db.begin_nested()
        try:
            if target_group:
                group_name = target_group.name
            else:
                group_name = str(row[col_map["group"]] or "").strip()
            lastname   = str(row[col_map["lastname"]] or "").strip()
            firstname  = str(row[col_map["firstname"]] or "").strip()
            if not group_name or not lastname or not firstname:
                skipped += 1
                sp.commit()
                continue
            _ph = "phone" in col_map and col_map["phone"] < len(row) and row[col_map["phone"]]
            phone = str(row[col_map["phone"]] if _ph else "").strip() or None
            _em = "email" in col_map and col_map["email"] < len(row) and row[col_map["email"]]
            email = str(row[col_map["email"]] if _em else "").strip().lower() or None

            # Mitglied suchen oder anlegen
            member = db.query(Member).filter(
                Member.org_id == org_id,
                Member.lastname == lastname,
                Member.firstname == firstname,
            ).first()
            if member:
                if phone:
                    member.phone = phone
                if email:
                    member.email = email
                member.active = True
                members_updated += 1
            else:
                member = Member(
                    lastname=lastname, firstname=firstname,
                    phone=phone, email=email,
                    org_id=org_id, active=True,
                )
                db.add(member)
                db.flush()
                members_created += 1

            # Gruppe suchen oder anlegen (Cache pro Import-Lauf)
            grp = group_cache.get(group_name)
            if grp is None:
                grp = db.query(SmsGroup).filter(
                    SmsGroup.org_id == org_id,
                    SmsGroup.name == group_name,
                ).first()
                if grp is None:
                    grp = SmsGroup(
                        org_id=org_id, name=group_name,
                        display_order=0, created_at=datetime.now(UTC),
                    )
                    db.add(grp)
                    db.flush()
                    groups_created += 1
                group_cache[group_name] = grp

            # Mitgliedschaft idempotent anlegen
            already = db.query(SmsGroupMember).filter_by(
                sms_group_id=grp.id, member_id=member.id
            ).first()
            if not already:
                db.add(SmsGroupMember(sms_group_id=grp.id, member_id=member.id))
                memberships_added += 1

            sp.commit()
        except Exception:
            sp.rollback()
            row_errors += 1

    if members_created or members_updated or groups_created or memberships_added:
        db.commit()
        write_audit(db, "admin.sms_group.excel_import", org_id=org_id, user_id=user.id,
                    payload={
                        "members_created": members_created,
                        "members_updated": members_updated,
                        "groups_created": groups_created,
                        "memberships_added": memberships_added,
                        "skipped": skipped,
                        "row_errors": row_errors,
                    })
    return RedirectResponse(
        f"/admin/gruppen?saved=1"
        f"&members_created={members_created}&members_updated={members_updated}"
        f"&groups_created={groups_created}&memberships_added={memberships_added}"
        f"&skipped={skipped}&row_errors={row_errors}",
        status_code=303,
    )


# ── Einsatzinfo-SMS-Konfiguration ─────────────────────────────────────────────

@router.get("/einsatzinfo-sms", response_class=HTMLResponse)
async def einsatzinfo_sms_page(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    user = request.state.user
    org_id = _require_org(user)

    org_settings = db.query(OrgSettings).filter(OrgSettings.org_id == org_id).first()
    alarm_types = (
        db.query(AlarmType)
        .filter(AlarmType.org_id == org_id)
        .order_by(AlarmType.category, AlarmType.code)
        .all()
    )
    groups = _sms_groups_for_org(db, org_id)
    members = _active_members(db, org_id)

    # Basis-Verteiler (alarm_type_id IS NULL)
    basis_recipients = (
        db.query(SmsEinsatzinfoRecipient)
        .filter(
            SmsEinsatzinfoRecipient.org_id == org_id,
            SmsEinsatzinfoRecipient.alarm_type_id.is_(None),
        )
        .all()
    )
    basis_group_ids = {r.group_id for r in basis_recipients if r.group_id}
    basis_member_ids = {r.member_id for r in basis_recipients if r.member_id}

    # Verteiler je Stichwort
    stichwort_recipients: dict[int, dict] = {}
    for at in alarm_types:
        recs = (
            db.query(SmsEinsatzinfoRecipient)
            .filter(
                SmsEinsatzinfoRecipient.org_id == org_id,
                SmsEinsatzinfoRecipient.alarm_type_id == at.id,
            )
            .all()
        )
        stichwort_recipients[at.id] = {
            "group_ids": {r.group_id for r in recs if r.group_id},
            "member_ids": {r.member_id for r in recs if r.member_id},
        }

    from app.services.sms_dispatch_service import default_einsatzinfo_template
    return templates.TemplateResponse(request, "admin/einsatzinfo_sms.html", {
        "user": user,
        "org_settings": org_settings,
        "alarm_types": alarm_types,
        "groups": groups,
        "members": members,
        "basis_group_ids": basis_group_ids,
        "basis_member_ids": basis_member_ids,
        "stichwort_recipients": stichwort_recipients,
        "default_template": default_einsatzinfo_template(),
        "saved": request.query_params.get("saved"),
    })


@router.post("/einsatzinfo-sms/einstellungen")
async def einsatzinfo_sms_save_settings(
    request: Request,
    enabled: bool = Form(False),
    send_exercise: bool = Form(False),
    template: str = Form(""),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Speichert Aktivierungsschalter und Org-Standard-Vorlage."""
    user = request.state.user
    org_id = _require_org(user)
    org_settings = db.query(OrgSettings).filter(OrgSettings.org_id == org_id).first()
    if not org_settings:
        raise HTTPException(status_code=404)
    org_settings.einsatzinfo_sms_enabled = enabled
    org_settings.einsatzinfo_sms_send_exercise = send_exercise
    org_settings.einsatzinfo_sms_template = template.strip() or None
    write_audit(db, "admin.einsatzinfo_sms.settings_saved", org_id=org_id, user_id=user.id,
                payload={"enabled": enabled, "send_exercise": send_exercise})
    db.commit()
    return RedirectResponse("/admin/einsatzinfo-sms?saved=1", status_code=303)


@router.post("/einsatzinfo-sms/basis-verteiler")
async def einsatzinfo_sms_save_basis(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Speichert den Basis-Verteiler (gilt fuer alle Stichworte, alarm_type_id=NULL)."""
    user = request.state.user
    org_id = _require_org(user)

    form = await request.form()
    group_ids = {int(v) for k, v in form.multi_items() if k == "group_id"}
    member_ids = {int(v) for k, v in form.multi_items() if k == "member_id"}

    # Basis-Eintraege loeschen und neu anlegen
    db.query(SmsEinsatzinfoRecipient).filter(
        SmsEinsatzinfoRecipient.org_id == org_id,
        SmsEinsatzinfoRecipient.alarm_type_id.is_(None),
    ).delete()

    for gid in group_ids:
        db.add(SmsEinsatzinfoRecipient(org_id=org_id, alarm_type_id=None, group_id=gid))
    for mid in member_ids:
        db.add(SmsEinsatzinfoRecipient(org_id=org_id, alarm_type_id=None, member_id=mid))

    write_audit(db, "admin.einsatzinfo_sms.basis_saved", org_id=org_id, user_id=user.id,
                payload={"groups": len(group_ids), "members": len(member_ids)})
    db.commit()
    return RedirectResponse("/admin/einsatzinfo-sms?saved=1", status_code=303)


@router.post("/einsatzinfo-sms/stichwort/{alarm_type_id}")
async def einsatzinfo_sms_save_stichwort(
    alarm_type_id: int,
    request: Request,
    template_override: str = Form(""),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Speichert Vorlagen-Override und Verteiler fuer ein einzelnes Stichwort."""
    user = request.state.user
    org_id = _require_org(user)

    at = db.get(AlarmType, alarm_type_id)
    if not at or at.org_id != org_id:
        raise HTTPException(status_code=404)

    # Vorlage am AlarmType speichern
    at.einsatzinfo_sms_template = template_override.strip() or None

    # Empfaenger-Eintraege ersetzen
    form = await request.form()
    group_ids = {int(v) for k, v in form.multi_items() if k == "group_id"}
    member_ids = {int(v) for k, v in form.multi_items() if k == "member_id"}

    db.query(SmsEinsatzinfoRecipient).filter(
        SmsEinsatzinfoRecipient.org_id == org_id,
        SmsEinsatzinfoRecipient.alarm_type_id == alarm_type_id,
    ).delete()

    for gid in group_ids:
        db.add(SmsEinsatzinfoRecipient(org_id=org_id, alarm_type_id=alarm_type_id, group_id=gid))
    for mid in member_ids:
        db.add(SmsEinsatzinfoRecipient(org_id=org_id, alarm_type_id=alarm_type_id, member_id=mid))

    write_audit(db, "admin.einsatzinfo_sms.stichwort_saved", org_id=org_id, user_id=user.id,
                entity_type="alarm_type", entity_id=alarm_type_id,
                payload={"groups": len(group_ids), "members": len(member_ids)})
    db.commit()
    return RedirectResponse(f"/admin/einsatzinfo-sms?saved=1#stichwort-{alarm_type_id}", status_code=303)


# ── Manueller SMS-Versand ─────────────────────────────────────────────────────

@router.get("/sms-senden", response_class=HTMLResponse)
async def sms_send_page(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    user = request.state.user
    org_id = _require_org(user)

    from app.routers.ws import is_sms_gateway_connected
    groups = _sms_groups_for_org(db, org_id)
    members = _active_members(db, org_id)
    sms_logs = (
        db.query(SmsLog)
        .filter(SmsLog.org_id == org_id)
        .order_by(SmsLog.sent_at.desc())
        .limit(30)
        .all()
    )
    return templates.TemplateResponse(request, "admin/sms_send.html", {
        "user": user,
        "groups": groups,
        "members": members,
        "sms_logs": sms_logs,
        "gateway_connected": is_sms_gateway_connected(org_id),
        "sent": request.query_params.get("sent"),
        "error": request.query_params.get("error"),
    })


@router.post("/sms-senden/senden")
async def sms_send_execute(
    request: Request,
    text: str = Form(...),
    target_type: str = Form("group"),  # "group" | "member" | "adhoc"
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Sendet eine manuelle SMS an Gruppen, Mitglieder oder Ad-hoc-Nummer."""
    user = request.state.user
    org_id = _require_org(user)
    text = text.strip()
    if not text:
        return RedirectResponse("/admin/sms-senden?error=empty", status_code=303)

    from app.routers.ws import is_sms_gateway_connected
    if not is_sms_gateway_connected(org_id):
        return RedirectResponse("/admin/sms-senden?error=no_gateway", status_code=303)

    form = await request.form()

    # Empfaenger aus Formular zusammenstellen
    import re as _re
    _strip_re = _re.compile(r"[\s\-\(\)]")

    phones: dict[str, str] = {}  # normalisierte Nummer → Anzeigename

    if target_type == "adhoc":
        adhoc_raw = (form.get("adhoc_number") or "").strip()
        if adhoc_raw:
            norm = _strip_re.sub("", adhoc_raw)
            phones[norm] = adhoc_raw
    else:
        group_ids = [int(v) for k, v in form.multi_items() if k == "group_id"]
        member_ids = [int(v) for k, v in form.multi_items() if k == "member_id"]

        # Gruppen expandieren
        if group_ids:
            groups = db.query(SmsGroup).filter(
                SmsGroup.id.in_(group_ids), SmsGroup.org_id == org_id
            ).all()
            for grp in groups:
                for gm in grp.members:
                    m = gm.member
                    if m and m.active and m.phone:
                        norm = _strip_re.sub("", m.phone.strip())
                        if norm:
                            phones[norm] = m.full_name

        # Einzelne Mitglieder
        if member_ids:
            mems = db.query(Member).filter(
                Member.id.in_(member_ids), Member.org_id == org_id, Member.active.is_(True)
            ).all()
            for m in mems:
                if m.phone:
                    norm = _strip_re.sub("", m.phone.strip())
                    if norm:
                        phones[norm] = m.full_name

    if not phones:
        return RedirectResponse("/admin/sms-senden?error=no_recipients", status_code=303)

    from app.services.sms_dispatch_service import send_bulk
    from app.core.audit import write_audit

    jobs = [(phone, text) for phone in phones]
    total, success = await send_bulk(org_id, jobs)

    # Protokollieren
    log_entry = SmsLog(
        org_id=org_id,
        sent_at=datetime.now(UTC),
        source="manual",
        alarm_type_code=None,
        text=text,
        recipient_count=total,
        success_count=success,
        triggered_by_user_id=user.id,
    )
    db.add(log_entry)
    write_audit(db, "admin.sms.manual_send", org_id=org_id, user_id=user.id,
                payload={"recipient_count": total, "success_count": success, "target_type": target_type})
    db.commit()

    return RedirectResponse(f"/admin/sms-senden?sent={success}", status_code=303)
