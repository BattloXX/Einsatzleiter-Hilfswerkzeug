"""Termine (Übung/Veranstaltung) + Teilnehmerlisten."""
from __future__ import annotations

import io
import logging
from datetime import UTC, datetime

from app.core.timezones import format_local_datetime, local_input_to_utc

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy import select as sa_select
from sqlalchemy.orm import Session

from app.core.dependencies import CurrentOrgId
from app.core.permissions import has_role
from app.core.templating import templates
from app.db import get_db
from app.models.master import Member, VehicleMaster
from app.models.sms import SmsGroup, SmsGroupMember
from app.models.teilnahme import Funktion, Teilnahme, Termin
from app.models.user import DeviceToken

router = APIRouter()
logger = logging.getLogger("einsatzleiter.termin")

_BEZUG_TYPEN = {"einsatz", "uebung", "veranstaltung"}

_STANDARD_FUNKTIONEN_NAMES = [
    "Einsatzleiter",
    "Gruppenkommandant",
    "Maschinist",
    "Atemschutzträger",
    "Melder",
    "Fahrer",
    "Truppführer",
    "Truppmann",
    "Sanitäter",
    "Sonstige",
]


def _require_login(request: Request):
    user = getattr(request.state, "user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Nicht angemeldet")
    return user


def _can_edit(user) -> bool:
    return has_role(user, "incident_leader", "recorder")


def _termin_or_404(termin_id: int, db: Session) -> Termin:
    t = db.get(Termin, termin_id)
    if not t:
        raise HTTPException(status_code=404, detail="Termin nicht gefunden")
    return t


def _teilnahme_or_404(teilnahme_id: int, db: Session) -> Teilnahme:
    t = db.get(Teilnahme, teilnahme_id)
    if not t:
        raise HTTPException(status_code=404, detail="Teilnahme nicht gefunden")
    return t


def _ensure_standard_funktionen(db: Session, org_id: int) -> None:
    """Legt Standardfunktionen an, falls noch keine für diese Org vorhanden sind."""
    count = db.query(Funktion).count()
    if count == 0:
        for i, name in enumerate(_STANDARD_FUNKTIONEN_NAMES, start=1):
            db.add(Funktion(org_id=org_id, name=name, sortierung=i, aktiv=True))
        db.flush()


def _load_funktionen(db: Session, org_id: int | None) -> list[Funktion]:
    if org_id:
        _ensure_standard_funktionen(db, org_id)
    return db.query(Funktion).filter(Funktion.aktiv == True).order_by(Funktion.sortierung, Funktion.name).all()  # noqa: E712


def _load_fahrzeuge(db: Session) -> list[VehicleMaster]:
    return db.query(VehicleMaster).filter(
        VehicleMaster.active == True,  # noqa: E712
        VehicleMaster.deleted == False,  # noqa: E712
    ).order_by(VehicleMaster.display_order, VehicleMaster.code).all()


# ── Termin-Liste ──────────────────────────────────────────────────────────────

@router.get("/termine", response_class=HTMLResponse)
async def termin_liste(
    request: Request,
    typ: str | None = None,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    termine = []
    incidents = []
    alle_eintraege = []
    if typ == "einsatz":
        from app.models.incident import Incident
        from app.routers.ui_archive import _scoped_incidents_query
        incidents = _scoped_incidents_query(db, user).order_by(Incident.started_at.desc()).limit(200).all()
    elif typ in ("uebung", "veranstaltung"):
        termine = db.query(Termin).filter(Termin.typ == typ).order_by(Termin.beginn.desc()).all()
    else:
        # "Alle": Termine und Einsätze gemeinsam nach Datum sortiert, damit Einsätze
        # direkt aus der Übersicht heraus für die Teilnahme-Erfassung ausgewählt
        # werden können (statt nur über den separaten "Einsätze"-Tab).
        from app.models.incident import Incident
        from app.routers.ui_archive import _scoped_incidents_query
        termine = db.query(Termin).order_by(Termin.beginn.desc()).all()
        incidents = _scoped_incidents_query(db, user).order_by(Incident.started_at.desc()).limit(200).all()
        alle_eintraege = sorted(
            [{"art": "termin", "obj": t, "dt": t.beginn} for t in termine]
            + [{"art": "einsatz", "obj": i, "dt": i.started_at} for i in incidents],
            key=lambda entry: entry["dt"] or datetime.min,
            reverse=True,
        )
    return templates.TemplateResponse(request, "termin/liste.html", {
        "user": user,
        "termine": termine,
        "incidents": incidents,
        "alle_eintraege": alle_eintraege,
        "filter_typ": typ or "",
        "can_edit": _can_edit(user),
    })


# ── Termin anlegen ────────────────────────────────────────────────────────────

@router.get("/termine/neu", response_class=HTMLResponse)
async def termin_neu_formular(
    request: Request,
    typ: str = "uebung",
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    return templates.TemplateResponse(request, "termin/formular.html", {
        "user": user,
        "termin": None,
        "default_typ": typ if typ in ("uebung", "veranstaltung") else "uebung",
    })


@router.post("/termine", response_class=HTMLResponse)
async def termin_anlegen(
    request: Request,
    typ: str = Form(...),
    titel: str = Form(...),
    beschreibung: str = Form(""),
    ort: str = Form(""),
    beginn: str = Form(...),
    ende: str = Form(""),
    ganztaegig: str = Form(""),
    status: str = Form("geplant"),
    db: Session = Depends(get_db),
    org_id: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    if typ not in ("uebung", "veranstaltung"):
        raise HTTPException(status_code=422, detail="Ungültiger Typ")
    beginn_dt = local_input_to_utc(beginn, user.org)
    if beginn_dt is None:
        raise HTTPException(status_code=422, detail="Ungültiges Datum")
    ende_dt = local_input_to_utc(ende, user.org) if ende else None
    t = Termin(
        org_id=org_id,
        typ=typ,
        titel=titel.strip(),
        beschreibung=beschreibung.strip() or None,
        ort=ort.strip() or None,
        beginn=beginn_dt,
        ende=ende_dt,
        ganztaegig=bool(ganztaegig),
        status=status if status in ("geplant", "laufend", "abgeschlossen", "abgesagt") else "geplant",
        erstellt_von=user.id,
        erstellt_am=datetime.now(UTC),
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return RedirectResponse(f"/termine/{t.id}", status_code=303)


# ── Termin-Detail ─────────────────────────────────────────────────────────────

@router.get("/termine/{termin_id}", response_class=HTMLResponse)
async def termin_detail(
    request: Request,
    termin_id: int,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    termin = _termin_or_404(termin_id, db)
    teilnahmen = _lade_teilnahmen(db, "uebung" if termin.typ == "uebung" else "veranstaltung", termin_id)
    funktionen = _load_funktionen(db, getattr(user, "org_id", None))
    fahrzeuge = _load_fahrzeuge(db)
    return templates.TemplateResponse(request, "termin/detail.html", {
        "user": user,
        "termin": termin,
        "teilnahmen": teilnahmen,
        "funktionen": funktionen,
        "fahrzeuge": fahrzeuge,
        "can_edit": _can_edit(user),
    })


# ── Termin bearbeiten ─────────────────────────────────────────────────────────

@router.get("/termine/{termin_id}/bearbeiten", response_class=HTMLResponse)
async def termin_bearbeiten_formular(
    request: Request,
    termin_id: int,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    termin = _termin_or_404(termin_id, db)
    return templates.TemplateResponse(request, "termin/formular.html", {
        "user": user,
        "termin": termin,
        "default_typ": termin.typ,
    })


@router.post("/termine/{termin_id}", response_class=HTMLResponse)
async def termin_speichern(
    request: Request,
    termin_id: int,
    typ: str = Form(...),
    titel: str = Form(...),
    beschreibung: str = Form(""),
    ort: str = Form(""),
    beginn: str = Form(...),
    ende: str = Form(""),
    ganztaegig: str = Form(""),
    status: str = Form("geplant"),
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    termin = _termin_or_404(termin_id, db)
    beginn_dt = local_input_to_utc(beginn, user.org)
    if beginn_dt is None:
        raise HTTPException(status_code=422, detail="Ungültiges Datum")
    ende_dt = local_input_to_utc(ende, user.org) if ende else None
    termin.typ = typ if typ in ("uebung", "veranstaltung") else termin.typ
    termin.titel = titel.strip()
    termin.beschreibung = beschreibung.strip() or None
    termin.ort = ort.strip() or None
    termin.beginn = beginn_dt
    termin.ende = ende_dt
    termin.ganztaegig = bool(ganztaegig)
    termin.status = status if status in ("geplant", "laufend", "abgeschlossen", "abgesagt") else termin.status
    db.commit()
    return RedirectResponse(f"/termine/{termin_id}", status_code=303)


@router.post("/termine/{termin_id}/loeschen")
async def termin_loeschen(
    request: Request,
    termin_id: int,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    termin = _termin_or_404(termin_id, db)
    db.query(Teilnahme).filter(
        Teilnahme.bezug_typ == termin.typ,
        Teilnahme.bezug_id == termin_id,
    ).delete()
    db.delete(termin)
    db.commit()
    return RedirectResponse("/termine", status_code=303)


# ── Teilnehmer-Komponente ─────────────────────────────────────────────────────

def _lade_teilnahmen(db: Session, bezug_typ: str, bezug_id: int) -> list[Teilnahme]:
    return db.query(Teilnahme).filter(
        Teilnahme.bezug_typ == bezug_typ,
        Teilnahme.bezug_id == bezug_id,
    ).order_by(Teilnahme.hinzugefuegt_am).all()


@router.get("/teilnahme/{bezug_typ}/{bezug_id}/liste", response_class=HTMLResponse)
async def teilnahme_liste(
    request: Request,
    bezug_typ: str,
    bezug_id: int,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if bezug_typ not in _BEZUG_TYPEN:
        raise HTTPException(status_code=422, detail="Ungültiger Bezug-Typ")
    teilnahmen = _lade_teilnahmen(db, bezug_typ, bezug_id)
    funktionen = _load_funktionen(db, getattr(user, "org_id", None))
    fahrzeuge = _load_fahrzeuge(db)
    return templates.TemplateResponse(request, "termin/_teilnahme_liste.html", {
        "user": user,
        "teilnahmen": teilnahmen,
        "bezug_typ": bezug_typ,
        "bezug_id": bezug_id,
        "funktionen": funktionen,
        "fahrzeuge": fahrzeuge,
        "can_edit": _can_edit(user),
    })


@router.get("/teilnahme/{bezug_typ}/{bezug_id}/auswahl", response_class=HTMLResponse)
async def teilnahme_auswahl(
    request: Request,
    bezug_typ: str,
    bezug_id: int,
    suche: str = "",
    nur_aktiv: str = "1",
    container_id: str = "teilnahme-liste-container",
    dialog_id: str = "auswahl-dialog",
    liste_id: str = "auswahl-liste",
    gruppe_ids: list[int] = Query(default=[]),
    db: Session = Depends(get_db),
    org_id: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    if bezug_typ not in _BEZUG_TYPEN:
        raise HTTPException(status_code=422, detail="Ungültiger Bezug-Typ")

    # Gruppen für diese Org laden (via TenantScoped automatisch gefiltert)
    gruppen = db.query(SmsGroup).order_by(SmsGroup.display_order, SmsGroup.name).all()

    # Mitglied-ID → Gruppen-Namen-Mapping für Badges im Template
    member_to_gruppen: dict[int, list[str]] = {}
    for g in gruppen:
        for gm in g.members:
            member_to_gruppen.setdefault(gm.member_id, []).append(g.name)

    # Bereits registrierte Mitglieder-IDs für diesen Bezug
    existing_subq = sa_select(Teilnahme.mitglied_id).where(
        Teilnahme.org_id == org_id,
        Teilnahme.bezug_typ == bezug_typ,
        Teilnahme.bezug_id == bezug_id,
        Teilnahme.mitglied_id.is_not(None),
    )

    q = db.query(Member).filter(Member.id.not_in(existing_subq))
    if nur_aktiv != "0":
        q = q.filter(Member.active == True)  # noqa: E712
    if suche.strip():
        s = f"%{suche.strip()}%"
        from sqlalchemy import or_
        q = q.filter(or_(Member.firstname.ilike(s), Member.lastname.ilike(s)))
    if gruppe_ids:
        mitglieder_in_gruppe = sa_select(SmsGroupMember.member_id).where(
            SmsGroupMember.sms_group_id.in_(gruppe_ids)
        )
        q = q.filter(Member.id.in_(mitglieder_in_gruppe))
    verfuegbare = q.order_by(Member.lastname, Member.firstname).all()

    return templates.TemplateResponse(request, "termin/_teilnahme_auswahl.html", {
        "user": user,
        "verfuegbare": verfuegbare,
        "bezug_typ": bezug_typ,
        "bezug_id": bezug_id,
        "suche": suche,
        "nur_aktiv": nur_aktiv,
        "container_id": container_id,
        "dialog_id": dialog_id,
        "liste_id": liste_id,
        "gruppen": gruppen,
        "selected_gruppe_ids": set(gruppe_ids),
        "member_to_gruppen": member_to_gruppen,
    })


@router.post("/teilnahme/{bezug_typ}/{bezug_id}", response_class=HTMLResponse)
async def teilnahme_hinzufuegen(
    request: Request,
    bezug_typ: str,
    bezug_id: int,
    db: Session = Depends(get_db),
    org_id: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    if bezug_typ not in _BEZUG_TYPEN:
        raise HTTPException(status_code=422, detail="Ungültiger Bezug-Typ")

    form = await request.form()
    mitglied_ids = form.getlist("mitglied_id")
    freitext = (form.get("freitext_name") or "").strip()  # type: ignore[union-attr]

    # Fahrzeug vom Tablet-Device automatisch ermitteln (nur bei Einsatz)
    device_fahrzeug_id: int | None = None
    if bezug_typ == "einsatz" and getattr(user, "is_device", False):
        device_token = db.query(DeviceToken).filter(
            DeviceToken.user_id == user.id,
            DeviceToken.revoked_at == None,  # noqa: E711
        ).first()
        if device_token and device_token.vehicle_master_id:
            device_fahrzeug_id = device_token.vehicle_master_id

    # Mehrfach-Insert Mitglieder
    for mid in mitglied_ids:
        try:
            mid_int = int(mid)  # type: ignore[arg-type]
        except (ValueError, TypeError):
            continue
        status = (form.get(f"s_{mid_int}") or "").strip()  # type: ignore[union-attr]
        exists = db.query(Teilnahme).filter(
            Teilnahme.org_id == org_id,
            Teilnahme.bezug_typ == bezug_typ,
            Teilnahme.bezug_id == bezug_id,
            Teilnahme.mitglied_id == mid_int,
        ).first()
        if not exists:
            db.add(Teilnahme(
                org_id=org_id,
                bezug_typ=bezug_typ,
                bezug_id=bezug_id,
                mitglied_id=mid_int,
                ausgerueckt=(status == "teilgenommen"),
                entschuldigt=(status == "entschuldigt"),
                fahrzeug_id=device_fahrzeug_id,
                hinzugefuegt_von=user.id,
                hinzugefuegt_am=datetime.now(UTC),
            ))

    # Freitext-Eintrag
    if freitext:
        db.add(Teilnahme(
            org_id=org_id,
            bezug_typ=bezug_typ,
            bezug_id=bezug_id,
            mitglied_id=None,
            freitext_name=freitext,
            fahrzeug_id=device_fahrzeug_id,
            hinzugefuegt_von=user.id,
            hinzugefuegt_am=datetime.now(UTC),
        ))

    db.commit()
    teilnahmen = _lade_teilnahmen(db, bezug_typ, bezug_id)
    funktionen = _load_funktionen(db, org_id)
    fahrzeuge = _load_fahrzeuge(db)
    return templates.TemplateResponse(request, "termin/_teilnahme_liste.html", {
        "user": user,
        "teilnahmen": teilnahmen,
        "bezug_typ": bezug_typ,
        "bezug_id": bezug_id,
        "funktionen": funktionen,
        "fahrzeuge": fahrzeuge,
        "can_edit": _can_edit(user),
    })


@router.patch("/teilnahme/{teilnahme_id}", response_class=HTMLResponse)
async def teilnahme_bearbeiten(
    request: Request,
    teilnahme_id: int,
    db: Session = Depends(get_db),
    org_id: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    teilnahme = _teilnahme_or_404(teilnahme_id, db)

    form = await request.form()
    if "funktion_id" in form:
        v = form.get("funktion_id")
        teilnahme.funktion_id = int(v) if v else None  # type: ignore[arg-type]
    if "fahrzeug_id" in form:
        v = form.get("fahrzeug_id")
        teilnahme.fahrzeug_id = int(v) if v else None  # type: ignore[arg-type]
    if "notiz" in form:
        teilnahme.notiz = (form.get("notiz") or "").strip() or None  # type: ignore[union-attr]
    if "ausgerueckt" in form:
        teilnahme.ausgerueckt = form.get("ausgerueckt") == "1"
    if "entschuldigt" in form:
        teilnahme.entschuldigt = form.get("entschuldigt") == "1"
    db.commit()

    # Partial-Response: aktualisierte Zeile
    teilnahmen = _lade_teilnahmen(db, teilnahme.bezug_typ, teilnahme.bezug_id)
    funktionen = _load_funktionen(db, org_id)
    fahrzeuge = _load_fahrzeuge(db)
    return templates.TemplateResponse(request, "termin/_teilnahme_liste.html", {
        "user": user,
        "teilnahmen": teilnahmen,
        "bezug_typ": teilnahme.bezug_typ,
        "bezug_id": teilnahme.bezug_id,
        "funktionen": funktionen,
        "fahrzeuge": fahrzeuge,
        "can_edit": _can_edit(user),
    })


@router.delete("/teilnahme/{teilnahme_id}", response_class=HTMLResponse)
async def teilnahme_entfernen(
    request: Request,
    teilnahme_id: int,
    db: Session = Depends(get_db),
    org_id: CurrentOrgId = None,
):
    user = _require_login(request)
    if not _can_edit(user):
        raise HTTPException(status_code=403, detail="Keine Berechtigung")
    teilnahme = _teilnahme_or_404(teilnahme_id, db)
    bezug_typ = teilnahme.bezug_typ
    bezug_id = teilnahme.bezug_id
    db.delete(teilnahme)
    db.commit()

    teilnahmen = _lade_teilnahmen(db, bezug_typ, bezug_id)
    funktionen = _load_funktionen(db, org_id)
    fahrzeuge = _load_fahrzeuge(db)
    return templates.TemplateResponse(request, "termin/_teilnahme_liste.html", {
        "user": user,
        "teilnahmen": teilnahmen,
        "bezug_typ": bezug_typ,
        "bezug_id": bezug_id,
        "funktionen": funktionen,
        "fahrzeuge": fahrzeuge,
        "can_edit": _can_edit(user),
    })


# ── Export ────────────────────────────────────────────────────────────────────

@router.get("/teilnahme/{bezug_typ}/{bezug_id}/druck", response_class=HTMLResponse)
async def teilnahme_druck(
    request: Request,
    bezug_typ: str,
    bezug_id: int,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if bezug_typ not in _BEZUG_TYPEN:
        raise HTTPException(status_code=422, detail="Ungültiger Bezug-Typ")
    teilnahmen = _lade_teilnahmen(db, bezug_typ, bezug_id)
    titel, beginn_dt, ort_str = _bezug_meta(db, bezug_typ, bezug_id)
    return templates.TemplateResponse(request, "termin/druck.html", {
        "user": user,
        "teilnahmen": teilnahmen,
        "bezug_typ": bezug_typ,
        "bezug_id": bezug_id,
        "titel": titel,
        "beginn": beginn_dt,
        "ort": ort_str,
        "now": datetime.now(UTC),
    })


@router.get("/teilnahme/{bezug_typ}/{bezug_id}/export.pdf")
async def teilnahme_export_pdf(
    request: Request,
    bezug_typ: str,
    bezug_id: int,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if bezug_typ not in _BEZUG_TYPEN:
        raise HTTPException(status_code=422, detail="Ungültiger Bezug-Typ")
    teilnahmen = _lade_teilnahmen(db, bezug_typ, bezug_id)
    titel, beginn_dt, ort_str = _bezug_meta(db, bezug_typ, bezug_id)

    from app.services.pdf_service import render_teilnahme_pdf
    pdf_bytes = render_teilnahme_pdf(
        teilnahmen=teilnahmen,
        bezug_typ=bezug_typ,
        titel=titel,
        beginn=beginn_dt,
        ort=ort_str,
        user=user,
        base_url=str(request.base_url),
    )
    safe_titel = titel.replace(" ", "_")[:40] if titel else bezug_typ
    filename = f"Teilnahme_{safe_titel}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/teilnahme/{bezug_typ}/{bezug_id}/export.xlsx")
async def teilnahme_export_xlsx(
    request: Request,
    bezug_typ: str,
    bezug_id: int,
    db: Session = Depends(get_db),
    _: CurrentOrgId = None,
):
    user = _require_login(request)
    if bezug_typ not in _BEZUG_TYPEN:
        raise HTTPException(status_code=422, detail="Ungültiger Bezug-Typ")
    teilnahmen = _lade_teilnahmen(db, bezug_typ, bezug_id)
    titel, beginn_dt, ort_str = _bezug_meta(db, bezug_typ, bezug_id)

    xlsx_bytes = _build_xlsx(teilnahmen, bezug_typ, titel, beginn_dt, ort_str, org=user.org)
    safe_titel = titel.replace(" ", "_")[:40] if titel else bezug_typ
    filename = f"Teilnahme_{safe_titel}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _bezug_meta(db: Session, bezug_typ: str, bezug_id: int) -> tuple[str, datetime | None, str | None]:
    """Laedt Titel, Beginn und Ort abhängig vom Bezug-Typ."""
    if bezug_typ in ("uebung", "veranstaltung"):
        t = db.get(Termin, bezug_id)
        if t:
            return t.titel, t.beginn, t.ort
        return f"{bezug_typ.capitalize()} #{bezug_id}", None, None
    else:
        from app.models.incident import Incident
        inc = db.get(Incident, bezug_id)
        if inc:
            adresse = " ".join(filter(None, [inc.address_street, inc.address_no, inc.address_city]))
            return f"Einsatz #{inc.id} – {inc.alarm_type_code}", inc.started_at, adresse or None
        return f"Einsatz #{bezug_id}", None, None


def _build_xlsx(
    teilnahmen: list[Teilnahme],
    bezug_typ: str,
    titel: str,
    beginn: datetime | None,
    ort: str | None,
    org=None,
) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teilnehmerliste"

    is_einsatz = bezug_typ == "einsatz"

    # Kopfzeile
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="D42225")

    cols = ["Nr.", "Name", "Aktiv"]
    if is_einsatz:
        cols += ["Funktion", "Fahrzeug"]
    else:
        cols += ["Teilgenommen", "Entschuldigt"]
    cols += ["Notiz", "Hinzugefügt von", "Hinzugefügt am"]

    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # Daten
    for ri, t in enumerate(teilnahmen, start=2):
        row_data = [ri - 1, t.anzeige_name, "✓" if (t.mitglied and t.mitglied.active) else "–"]
        if is_einsatz:
            row_data += [
                t.funktion.name if t.funktion else "",
                t.fahrzeug.display_label if t.fahrzeug else "",
            ]
        else:
            row_data += [
                "✓" if t.ausgerueckt else "",
                "✓" if t.entschuldigt else "",
            ]
        row_data.append(t.notiz or "")
        row_data.append(t.hinzugefuegt_von_user.display_name if t.hinzugefuegt_von_user else "")
        row_data.append(format_local_datetime(t.hinzugefuegt_am, org))
        for ci, val in enumerate(row_data, start=1):
            ws.cell(row=ri, column=ci, value=val)

    # Autofilter + Spaltenbreiten
    ws.auto_filter.ref = ws.dimensions
    widths = [5, 28, 8]
    if is_einsatz:
        widths += [20, 18]
    else:
        widths += [14, 14]
    widths += [30, 22, 16]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Metadaten als Kommentar in Zelle A1
    ws["A1"].comment = None  # vorsorglich
    meta_parts = [titel]
    if beginn:
        meta_parts.append(format_local_datetime(beginn, org))
    if ort:
        meta_parts.append(ort)
    ws["A1"].comment = None  # keep it clean

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
